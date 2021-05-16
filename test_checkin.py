import cv2
import mediapipe as mp 
import pickle
import numpy as np
from tensorflow.keras.models import load_model
from keras_facenet import FaceNet
from sklearn import svm
from scipy.spatial.distance import cosine
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from pathlib import Path

filename = 'data/Models/Timekeeping.xlsx'
df = pd.read_excel(filename)

video_capture = cv2.VideoCapture(0)

mpFaceDetect = mp.solutions.face_detection
mpDraw = mp.solutions.drawing_utils

faceDetection = mpFaceDetect.FaceDetection(0.8)

SVM = pickle.load(open("Models/model_using_svm.sav", "rb"))

get_date = datetime.utcnow().strftime('%d/%m/%Y')

data_get_date = df[(df['date_logtime'] == get_date)]
user_in_data_get_date = data_get_date['user_id'].tolist()

def addDataExcel(data, filename):
    wb_obj = load_workbook(filename)
    sheet_obj = wb_obj.active
    row = sheet_obj.max_row
    column = sheet_obj.max_column
    for j in range(1, column+1):
        sheet_obj.cell(row = row + 1, column= j, value=data[j-1])

    wb_obj.save(filename)


model = FaceNet()

def fancyBox(img, bbox, l=30, t=2, rt=1):
    x, y, w, h = bbox
    x1, y1 = x + w, y + h
    # cv2.rectangle(img, bbox, (255, 0, 255), rt)
    # Top left x, y
    cv2.line(img, (x, y), (x + l, y), (255, 0, 255), t)
    cv2.line(img, (x, y), (x, y + l), (255, 0, 255), t)
    # Top right x, y
    cv2.line(img, (x1, y), (x1 - l, y), (255, 0, 255), t)
    cv2.line(img, (x1, y), (x1, y + l), (255, 0, 255), t)
    # Bottom left x, y
    cv2.line(img, (x, y1), (x + l, y1), (255, 0, 255), t)
    cv2.line(img, (x, y1), (x, y1 - l), (255, 0, 255), t)
    # Bottom right x, y
    cv2.line(img, (x1, y1), (x1 - l, y1), (255, 0, 255), t)
    cv2.line(img, (x1, y1), (x1, y1 - l), (255, 0, 255), t)

while True:
    _, img = video_capture.read()
    
    # img = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
    results = faceDetection.process(img)
    
    if results.detections:
        for id, detection in enumerate(results.detections):
            # mpDraw.draw_detection(img, detection)
            bboxC = detection.location_data.relative_bounding_box
            ih, iw, ic = img.shape
            bbox = int(bboxC.xmin * iw) - 20, int(bboxC.ymin * ih) - 45, int(bboxC.width * iw) + 40, int(bboxC.height * ih) + 50
            
            x, y, w, h = bbox
            roi_color = img[y:y+h, x:x+w]
            if roi_color.shape[0] != 0 and roi_color.shape[1] != 0:
                roi_color = cv2.resize(roi_color,(160,160))
                face_preprocess = cv2.cvtColor(roi_color, cv2.COLOR_BGR2RGB)
                face_preprocess = np.expand_dims(roi_color, axis = 0)

                face_feature = model.embeddings(face_preprocess)
                
                result = SVM.predict(np.array(face_feature))
                proba = SVM.predict_proba(np.array(face_feature))
                print(proba)
                fancyBox(img, bbox)
                cv2.putText(img, str(result[0]), (x - 10, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 2)
                print(result[0])
                
                if result[0] not in user_in_data_get_date:
                    timestr = datetime.now().strftime('%H-%M-%S-%f')
                    path_checkin = "data/timekeeping/" + str(datetime.now().strftime('%d-%m-%Y')) + '/' + str(result[0]) + '/'
                    Path(path_checkin).mkdir(parents=True, exist_ok=True)

                    cv2.imwrite(path_checkin + str(timestr) + ".jpg", roi_color)

                    data = [
                                datetime.now().strftime('%d/%m/%Y'), 
                                result[0],
                                datetime.now().strftime('%H:%M:%S'), 
                                '', 
                                path_checkin + str(timestr) + ".jpg", 
                                ''
                            ]

                    addDataExcel(data, filename)
                    user_in_data_get_date.append(result[0])
                
    cv2.imshow('Video', img)
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break
    
video_capture.release()
cv2.destroyAllWindows()




