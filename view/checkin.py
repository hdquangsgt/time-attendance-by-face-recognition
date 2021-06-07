import cv2
import mediapipe as mp 
import pickle
import numpy as np
from tensorflow.keras.models import load_model
from keras_facenet import FaceNet
from scipy.spatial.distance import cosine
import pandas as pd
from datetime import datetime
from time import sleep
from openpyxl import load_workbook
from pathlib import Path
import pyttsx3
import os

class CheckIn(object):
    def __init__(self, *args):
        engine = pyttsx3.init()
        voices = engine.getProperty("voices")

        filename = 'data/Models/Timekeeping.xlsx'
        df = pd.read_excel(filename)
        video_capture = cv2.VideoCapture(0)
        mpFaceDetect = mp.solutions.face_detection
        mpDraw = mp.solutions.drawing_utils

        faceDetection = mpFaceDetect.FaceDetection(0.7)

        with open("Models/data_embeddings.p", "rb") as f:
            data = pickle.load(f)

        get_date = datetime.utcnow().strftime('%d/%m/%Y')
        data_get_date = df[(df['date_logtime'] == get_date)]
        user_in_data_get_date = data_get_date['user_id'].tolist()

        model = FaceNet()
        
        while True:
            _, img = video_capture.read()
            # img = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
            results = faceDetection.process(img)
            
            if results.detections:
                for id, detection in enumerate(results.detections):
                    # mpDraw.draw_detection(img, detection)
                    bboxC = detection.location_data.relative_bounding_box
                    ih, iw, ic = img.shape
                    bbox = int(bboxC.xmin * iw) - 20, int(bboxC.ymin * ih) - 45, int(bboxC.width * iw) + 40, int(bboxC.height * ih) + 50

                    x, y, w, h = bbox
                    roi_color = img[y:y+h, x:x+w]
                    if roi_color.shape[0] != 0 and roi_color.shape[1] != 0:
                        roi_color = cv2.resize(roi_color,(160,160))
                        face_preprocess = cv2.cvtColor(roi_color, cv2.COLOR_BGR2RGB)
                        face_preprocess = np.expand_dims(roi_color, axis = 0)

                        face_feature = model.embeddings(face_preprocess)

                        distance = 99999
                        name = 'Unknow'
                        for db_name in data:
                            for encoding in data[db_name]:
                                dist = cosine(encoding, face_feature)
                                if dist < 0.3 and dist < distance:
                                    name = db_name
                                    distance = dist
                        cv2.putText(img, name, (x - 10, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.5, (0, 255, 0), 2)
                        print(distance, '->', name)
                        self.fancyBox(img, bbox)
                        if name != 'Unknow':
                            filenameEmployee = os.path.abspath('data/Models/Employee.xlsx')
                            employees = pd.read_excel(filenameEmployee)
                            listEmployee = employees.to_numpy().tolist()
                            employee = [x for x in listEmployee if x[1] == name]
                            if((name not in user_in_data_get_date) and employee != []):
                                nameEmployee = employee[0][2]
                                engine.setProperty("voice",voices[1].id)
                                engine.say("Xin chào " + nameEmployee)
                                engine.runAndWait()

                                timestr = datetime.now().strftime('%H-%M-%S-%f')

                                path_checkin = "data/timekeeping/checkin/" + str(datetime.now().strftime('%d-%m-%Y')) + '/' + str(name) + '/'
                                Path(path_checkin).mkdir(parents=True, exist_ok=True)

                                cv2.imwrite(path_checkin + str(timestr) + ".jpg", roi_color)

                                data_record = [
                                            datetime.now().strftime('%d/%m/%Y'), 
                                            name,
                                            datetime.now().strftime('%H:%M:%S'), 
                                            '', 
                                            path_checkin + str(timestr) + ".jpg", 
                                            ''
                                        ]
                                self.addDataExcel(data_record, filename)
                                user_in_data_get_date.append(name)

            cv2.imshow('Video', img)
            if cv2.waitKey(1) & 0xFF == ord('q'):
                break
        video_capture.release()
        cv2.destroyAllWindows()

    def addDataExcel(self, data, filename):
        wb_obj = load_workbook(filename)
        sheet_obj = wb_obj.active
        row = sheet_obj.max_row
        column = sheet_obj.max_column
        for j in range(1, column + 1):
            sheet_obj.cell(row = row + 1, column= j, value=data[j-1])

        wb_obj.save(filename)

    def fancyBox(self, img, bbox, l=30, t=2, rt=1):
        x, y, w, h = bbox
        x1, y1 = x + w, y + h
        # cv2.rectangle(img, bbox, (255, 0, 255), rt)
        # Top left x, y
        cv2.line(img, (x, y), (x + l, y), (255, 0, 255), t)
        cv2.line(img, (x, y), (x, y + l), (255, 0, 255), t)
        # Top right x, y
        cv2.line(img, (x1, y), (x1 - l, y), (255, 0, 255), t)
        cv2.line(img, (x1, y), (x1, y + l), (255, 0, 255), t)
        # Bottom left x, y
        cv2.line(img, (x, y1), (x + l, y1), (255, 0, 255), t)
        cv2.line(img, (x, y1), (x, y1 - l), (255, 0, 255), t)
        # Bottom right x, y
        cv2.line(img, (x1, y1), (x1 - l, y1), (255, 0, 255), t)
        cv2.line(img, (x1, y1), (x1, y1 - l), (255, 0, 255), t)

