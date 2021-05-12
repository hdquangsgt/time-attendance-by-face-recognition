from tkinter import *
from tkinter import messagebox
from PIL import ImageTk, Image
import os
import pandas as pd
from openpyxl import load_workbook
import re
from .datepicker import CustomDateEntry

class RegistryForm(object):
    def __init__(self, root):
        self.root = root
        self.root.title('Quản lý nhân viên')
        self.root.geometry('700x450+0+0')
        self.root.resizable(False, False)
        
        bg_color = '#990099'

        self.name_entry = ''
        self.email_entry = ''

        #========= All images =========#
        # imageBG = os.path.abspath('view/images/bg-login.png')
        # self.bg_icon = PhotoImage(file = imageBG)
        # bg_lbl = Label(self.root, image = self.bg_icon).place(x = 250, y = 0, relwidth = 1, relheight = 1)

        # icon_registry = os.path.abspath('view/images/registry.png')
        # self.icon_registry = PhotoImage(file=icon_registry)
        # registry_lbl = Label(self.root, image = self.icon_registry).place(x = 80, y = 80)

        registry_title = Label(self.root, 
                            text='Đăng ký nhân viên',
                            bg = bg_color,
                            fg = 'white',
                            compound = CENTER,
                            bd = 10,
                            relief = GROOVE,
                            font = ('time new roman', 24, 'bold'))
        registry_title.place(x = 0, y = 0, relwidth = 1)

        registry_frm = Frame(self.root, bg = 'white')
        registry_frm.place(x = 100, y = 100, width = 500, height = 300)

        #   Field name
        name_lbl = Label(registry_frm, text="Họ và tên", width = 20, compound = LEFT, font=("bold", 10))
        name_lbl.grid(row = 1, column = 1, padx = 10, pady = 10)

        self.name_entry = Entry(registry_frm, width = 40)
        self.name_entry.grid(row = 1, column = 2, padx = 10, pady = 10)

        #   Field birth
        birth_lbl = Label(registry_frm, text="Ngày sinh", width = 20, compound = LEFT, font=("bold", 10))
        birth_lbl.grid(row = 2, column = 1, padx = 10, pady = 10)

        self.birth_entry = CustomDateEntry(registry_frm, date_pattern='dd/MM/yyyy')
        self.birth_entry._set_text(self.birth_entry._date.strftime('%d/%m/%Y'))
        self.birth_entry.grid(row = 2, column = 2, padx = 10, pady = 10)
        
        #   Field email
        email_lbl = Label(registry_frm, text="Email", width = 20, compound = LEFT, font=("bold", 10))
        email_lbl.grid(row = 3, column = 1, padx = 10, pady = 10)

        self.email_entry = Entry(registry_frm, width = 40)
        self.email_entry.grid(row = 3, column = 2, padx = 10, pady = 10)

        Button(registry_frm, text='Trở về',width=20,bg='brown',fg='white',command=self.goToBack).place(x=50,y=250)
        Button(registry_frm, text='Đăng ký',width=20,bg='brown',fg='white',command=self.submit).place(x=300,y=250)
        
    def goToBack(self):
        from .employee import EmployeeGUI
        self.root.destroy()
        frame = Tk()
        employee = EmployeeGUI(frame)

    def addDataExcel(self, data, filename):
        wb_obj = load_workbook(filename)
        sheet_obj = wb_obj.active
        row = sheet_obj.max_row
        column = sheet_obj.max_column
        for j in range(1, column+1):
            sheet_obj.cell(row = row + 1, column= j, value=data[j-1])

        wb_obj.save(filename)

    def no_accent_vietnamese(self,s):
        s = re.sub('[áàảãạăắằẳẵặâấầẩẫậ]', 'a', s)
        s = re.sub('[ÁÀẢÃẠĂẮẰẲẴẶÂẤẦẨẪẬ]', 'A', s)
        s = re.sub('[éèẻẽẹêếềểễệ]', 'e', s)
        s = re.sub('[ÉÈẺẼẸÊẾỀỂỄỆ]', 'E', s)
        s = re.sub('[óòỏõọôốồổỗộơớờởỡợ]', 'o', s)
        s = re.sub('[ÓÒỎÕỌÔỐỒỔỖỘƠỚỜỞỠỢ]', 'O', s)
        s = re.sub('[íìỉĩị]', 'i', s)
        s = re.sub('[ÍÌỈĨỊ]', 'I', s)
        s = re.sub('[úùủũụưứừửữự]', 'u', s)
        s = re.sub('[ÚÙỦŨỤƯỨỪỬỮỰ]', 'U', s)
        s = re.sub('[ýỳỷỹỵ]', 'y', s)
        s = re.sub('[ÝỲỶỸỴ]', 'Y', s)
        s = re.sub('đ', 'd', s)
        s = re.sub('Đ', 'D', s)
        return s
    
    def getIdLast(self,filename):
        df = pd.read_excel(filename)
        return df.iloc[-1]['id']

    def submit(self):
        from .employee import EmployeeGUI
        filename = os.path.abspath('data/Models/Employee.xlsx')
        id = self.getIdLast(filename) + 1
        name = self.name_entry.get()
        
        validated = self.validate(name,self.email_entry.get())
        
        if (validated == ['','']):
            user_id = self.generateUserId(self.no_accent_vietnamese(name))
            data = [
                id,
                user_id,
                name,
                self.birth_entry.get_date(),
                self.email_entry.get(),
                ''
            ]
            self.addDataExcel(data, filename)
            self.createFolder(user_id)
            #   Exit
            employeeFrame = Tk()
            EmployeeGUI(employeeFrame)
            self.root.destroy()
        else:
            global errorFrame
            errorFrame = Tk()
            errorFrame.geometry('200x80')
            errorFrame.title('Kiểm tra dữ liệu nhập')

            if(validated[0]):
                Label(errorFrame, text = validated[0], fg = 'red').grid(row = 0, column = 0, padx = 20, pady = 5)
            elif(validated[1]):
                Label(errorFrame, text = validated[1], fg = 'red').grid(row = 1, column = 0, padx = 20, pady = 5)

            Button(errorFrame, text = 'Đã hiểu', command = self.deleteErrorFrame).grid(row = 3, column = 0, padx = 20, pady = 15)
            return

    def deleteErrorFrame(self):
        errorFrame.destroy()

    def createFolder(self,userId):
        parent_dir = os.path.abspath('data/face_train')
        path = os.path.join(parent_dir, userId)
        os.makedirs(path)

    def generateUserId(self,name):
        arrName = name.split()
        user_id = arrName[len(arrName) - 1]

        for i in range(len(arrName)):
            if (i < len(arrName) - 1):
                user_id = user_id + arrName[i][0]
        root_name = user_id
        count = 1
        filename = os.path.abspath('data/Models/Employee.xlsx')
        df = pd.read_excel(filename)
        df_rows = df.to_numpy().tolist()
        for row in df_rows:
            if root_name in row[1]:
                user_id = root_name + str(count)
                count = count + 1

        return user_id

    def validate(self,name,email):
        name_message = ''
        email_message = ''
        input = {
            'name' : name,
            'email': email
        }
        if(self.checkEmpty(input)):
            return self.checkEmpty(input)
        
        regex = '^[a-z0-9]+[\._]?[a-z0-9]+[@]\w+[.]\w{2,3}$'
        if(not re.search(regex,email)):   
            email_message = 'Định dạng email không hợp lệ'
        return [name_message,email_message]

    def checkEmpty(self,input):
        result = []
        if(len(input['name']) == 0):
            result.append('Tên không được để trống')

        if(len(input['email']) == 0):
            result.append('Email không được để trống')

        return result

